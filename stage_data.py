#!/usr/bin/env python3
"""Stage benchmark data: download KiTS, LiTS (MSD), PanTS and create Cruz_XXXXXXXX folders.

Layout created:
  MedAgentsBench/data/
    Kidney/
      public/Cruz_00000001/ct.nii.gz  ... Cruz_00000020/ct.nii.gz
      private/
        masks/Cruz_00000001/{organ.nii.gz, lesion.nii.gz} ...
        ground_truth.csv
    Liver/
      public/Cruz_00000021/ct.nii.gz  ... Cruz_00000040/ct.nii.gz
      private/...
    Pancreas/
      public/Cruz_00000041/ct.nii.gz  ... Cruz_00000060/ct.nii.gz
      private/...

Sources:
  Kidney  -> KiTS19 (DigitalOcean Spaces + GitHub segmentations)
  Liver   -> MSD Task03_Liver (HuggingFace)
  Pancreas -> PanTS (HuggingFace BodyMaps/PanTSMini)

All 60 cases are tumor-positive.
"""

import os
import sys
import json
import shutil
import tarfile
import tempfile
import subprocess
from pathlib import Path

import numpy as np
import nibabel as nib

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
TMP_DIR = ROOT / "tmp_staging"

# ── KiTS19 config ─────────────────────────────────────────────────────────
KITS_IMG_URL = "https://kits19.sfo2.digitaloceanspaces.com/master_{:05d}.nii.gz"
KITS_SEG_URL = "https://raw.githubusercontent.com/neheller/kits19/master/data/case_{:05d}/segmentation.nii.gz"
KITS_META_URL = "https://raw.githubusercontent.com/neheller/kits19/master/data/kits.json"
# KiTS label mapping: 0=background, 1=kidney, 2=tumor

# ── MSD Liver (LiTS equivalent) ───────────────────────────────────────────
MSD_LIVER_URL = "https://huggingface.co/datasets/Novel-BioMedAI/Medical_Segmentation_Decathlon/resolve/main/Task03_Liver.tar?download=true"
# MSD Liver label mapping: 0=background, 1=liver, 2=tumor

# ── PanTS config ──────────────────────────────────────────────────────────
PANTS_META_URL = "https://huggingface.co/datasets/BodyMaps/PanTSMini/resolve/main/metadata.xlsx?download=true"
PANTS_LABEL_URL = "http://www.cs.jhu.edu/~zongwei/dataset/PanTSMini_Label.tar.gz"
# PanTS images are in tar archives by range on HuggingFace
PANTS_IMG_URL_TMPL = "https://huggingface.co/datasets/BodyMaps/PanTSMini/resolve/main/PanTSMini_ImageTr_{start}_{end}.tar.gz?download=true"
# PanTS label mapping: 0=background, 1=pancreas, 2=tumor (PDAC), etc.


def run(cmd, **kwargs):
    print(f"  $ {cmd}", flush=True)
    subprocess.run(cmd, shell=True, check=True, **kwargs)


def download(url, dest, desc=None):
    """Download a file with wget (quiet, with retry)."""
    dest = Path(dest)
    if dest.exists() and dest.stat().st_size > 1000:
        print(f"  [skip] {dest.name} already exists ({dest.stat().st_size:,} bytes)")
        return
    dest.parent.mkdir(parents=True, exist_ok=True)
    label = desc or dest.name
    print(f"  Downloading {label} ...", flush=True)
    run(f'wget -q --tries=3 --timeout=120 -O "{dest}" "{url}"')
    print(f"  -> {dest.stat().st_size:,} bytes", flush=True)


def extract_organ_lesion(seg_path, organ_label, tumor_label, out_dir):
    """Extract binary organ and lesion masks from a multi-label segmentation."""
    seg = nib.load(str(seg_path))
    data = seg.get_fdata().astype(np.uint8)
    affine = seg.affine
    header = seg.header

    organ_mask = ((data == organ_label) | (data == tumor_label)).astype(np.uint8)
    lesion_mask = (data == tumor_label).astype(np.uint8)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    nib.save(nib.Nifti1Image(organ_mask, affine, header), str(out_dir / "organ.nii.gz"))
    nib.save(nib.Nifti1Image(lesion_mask, affine, header), str(out_dir / "lesion.nii.gz"))

    has_lesion = int(lesion_mask.sum() > 0)
    return has_lesion


# ═══════════════════════════════════════════════════════════════════════════
# KIDNEY (KiTS19)
# ═══════════════════════════════════════════════════════════════════════════
def stage_kidney():
    print("\n" + "=" * 60)
    print("STAGING KIDNEY (KiTS19) — 20 tumor-positive cases")
    print("=" * 60)

    organ_dir = DATA_DIR / "Kidney"
    pub_dir = organ_dir / "public"
    priv_dir = organ_dir / "private"
    masks_dir = priv_dir / "masks"
    tmp = TMP_DIR / "kits"
    tmp.mkdir(parents=True, exist_ok=True)

    # 1. Get metadata to find tumor-positive cases
    meta_file = tmp / "kits.json"
    download(KITS_META_URL, meta_file, "kits.json metadata")
    with open(meta_file) as f:
        meta = json.load(f)

    tumor_cases = [m for m in meta if m.get("malignant") is True]
    selected = tumor_cases[:20]
    print(f"  Selected {len(selected)} tumor-positive cases")

    # 2. Clone kits19 repo for segmentation labels (LFS-tracked)
    kits_repo = tmp / "kits19"
    if not (kits_repo / "data").exists():
        print("  Cloning kits19 repo (segmentation labels via Git LFS) ...")
        run(f'cd "{tmp}" && git lfs install 2>/dev/null; '
            f'git clone --depth 1 https://github.com/neheller/kits19.git 2>&1 | tail -3')
    else:
        print("  [skip] kits19 repo already cloned")

    gt_rows = []
    for idx, case in enumerate(selected, start=1):
        cruz_id = f"Cruz_{idx:08d}"
        case_num = int(case["case_id"].replace("case_", ""))
        print(f"\n  [{idx}/20] {case['case_id']} -> {cruz_id}", flush=True)

        # Download imaging from DigitalOcean
        img_tmp = tmp / f"case_{case_num:05d}_imaging.nii.gz"
        download(KITS_IMG_URL.format(case_num), img_tmp,
                 f"CT case_{case_num:05d} (~200MB)")

        # Segmentation from cloned repo
        seg_src = kits_repo / "data" / f"case_{case_num:05d}" / "segmentation.nii.gz"
        if not seg_src.exists() or seg_src.stat().st_size < 10000:
            print(f"    WARNING: segmentation not found or LFS stub for case_{case_num:05d}")
            continue

        # Copy CT to public
        ct_dest = pub_dir / cruz_id / "ct.nii.gz"
        ct_dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(img_tmp), str(ct_dest))

        # Extract organ (label 1+2) and lesion (label 2) masks
        has_lesion = extract_organ_lesion(
            seg_src, organ_label=1, tumor_label=2,
            out_dir=masks_dir / cruz_id
        )
        gt_rows.append(f"{cruz_id},kidney,{has_lesion}")

        # Remove downloaded CT to save space
        img_tmp.unlink(missing_ok=True)

    # Write ground truth CSV
    priv_dir.mkdir(parents=True, exist_ok=True)
    gt_path = priv_dir / "ground_truth.csv"
    with open(gt_path, "w") as f:
        f.write("patient_id,organ,lesion_present\n")
        f.write("\n".join(gt_rows) + "\n")
    print(f"\n  Wrote {gt_path}")


# ═══════════════════════════════════════════════════════════════════════════
# LIVER (MSD Task03_Liver / LiTS)
# ═══════════════════════════════════════════════════════════════════════════
def stage_liver():
    print("\n" + "=" * 60)
    print("STAGING LIVER (MSD Task03_Liver / LiTS) — 20 tumor-positive cases")
    print("=" * 60)

    organ_dir = DATA_DIR / "Liver"
    pub_dir = organ_dir / "public"
    priv_dir = organ_dir / "private"
    masks_dir = priv_dir / "masks"
    tmp = TMP_DIR / "liver"
    tmp.mkdir(parents=True, exist_ok=True)

    # 1. Download MSD Task03_Liver tar (~30GB)
    tar_path = tmp / "Task03_Liver.tar"
    download(MSD_LIVER_URL, tar_path, "Task03_Liver.tar (~30GB)")

    # 2. Extract to find tumor-positive cases
    extract_dir = tmp / "extracted"
    if not extract_dir.exists():
        print("  Extracting Task03_Liver.tar ...", flush=True)
        extract_dir.mkdir(parents=True, exist_ok=True)
        run(f'tar -xf "{tar_path}" -C "{extract_dir}"')

    # MSD structure: Task03_Liver/imagesTr/liver_NNN.nii.gz
    #                Task03_Liver/labelsTr/liver_NNN.nii.gz
    msd_root = extract_dir / "Task03_Liver"
    img_dir = msd_root / "imagesTr"
    lbl_dir = msd_root / "labelsTr"

    # 3. Scan for tumor-positive cases (label 2 = tumor)
    print("  Scanning for tumor-positive cases ...")
    label_files = sorted(lbl_dir.glob("liver_*.nii.gz"))
    tumor_cases = []
    for lf in label_files:
        seg = nib.load(str(lf))
        data = seg.get_fdata()
        if (data == 2).any():
            case_name = lf.stem.replace(".nii", "")
            tumor_cases.append(case_name)
            if len(tumor_cases) >= 20:
                break
        if len(tumor_cases) >= 20:
            break

    print(f"  Found {len(tumor_cases)} tumor-positive cases")

    gt_rows = []
    for idx, case_name in enumerate(tumor_cases, start=21):
        cruz_id = f"Cruz_{idx:08d}"
        print(f"  [{idx-20}/20] {case_name} -> {cruz_id}")

        img_src = img_dir / f"{case_name}.nii.gz"
        seg_src = lbl_dir / f"{case_name}.nii.gz"

        # Copy CT
        ct_dest = pub_dir / cruz_id / "ct.nii.gz"
        ct_dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(img_src), str(ct_dest))

        # Extract organ (label 1+2) and lesion (label 2) masks
        has_lesion = extract_organ_lesion(
            seg_src, organ_label=1, tumor_label=2,
            out_dir=masks_dir / cruz_id
        )
        gt_rows.append(f"{cruz_id},liver,{has_lesion}")

    # Write ground truth CSV
    priv_dir.mkdir(parents=True, exist_ok=True)
    gt_path = priv_dir / "ground_truth.csv"
    with open(gt_path, "w") as f:
        f.write("patient_id,organ,lesion_present\n")
        f.write("\n".join(gt_rows) + "\n")
    print(f"\n  Wrote {gt_path}")


# ═══════════════════════════════════════════════════════════════════════════
# PANCREAS (PanTS)
# ═══════════════════════════════════════════════════════════════════════════
def stage_pancreas():
    print("\n" + "=" * 60)
    print("STAGING PANCREAS (PanTS) — 20 tumor-positive cases")
    print("=" * 60)

    organ_dir = DATA_DIR / "Pancreas"
    pub_dir = organ_dir / "public"
    priv_dir = organ_dir / "private"
    masks_dir = priv_dir / "masks"
    tmp = TMP_DIR / "pants"
    tmp.mkdir(parents=True, exist_ok=True)

    # 1. Get metadata to find tumor-positive cases
    meta_file = tmp / "metadata.xlsx"
    download(PANTS_META_URL, meta_file)

    import openpyxl
    wb = openpyxl.load_workbook(str(meta_file))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    tumor_col = headers.index("tumor?")
    id_col = headers.index("PanTS ID")

    tumor_ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[tumor_col] == 1:
            tumor_ids.append(row[id_col])
        if len(tumor_ids) >= 20:
            break

    print(f"  Selected {len(tumor_ids)} tumor-positive cases: {tumor_ids[0]} .. {tumor_ids[-1]}")

    # 2. Figure out which tar archives we need (each covers 1000 cases)
    needed_archives = set()
    for pid in tumor_ids:
        num = int(pid.replace("PanTS_", ""))
        archive_idx = (num - 1) // 1000 + 1  # 1-based
        start = (archive_idx - 1) * 1000 + 1
        end = archive_idx * 1000
        needed_archives.add((archive_idx, f"{start:08d}", f"{end:08d}"))

    # 3. Download and extract needed image archives
    img_extract_dir = tmp / "ImageTr"
    img_extract_dir.mkdir(parents=True, exist_ok=True)

    for arc_idx, start, end in sorted(needed_archives):
        arc_name = f"PanTSMini_ImageTr_{start}_{end}.tar.gz"
        arc_path = tmp / arc_name
        url = PANTS_IMG_URL_TMPL.format(start=start, end=end)
        download(url, arc_path)
        # Extract only the cases we need
        print(f"  Extracting {arc_name} ...")
        run(f'tar -xzf "{arc_path}" -C "{img_extract_dir}"')

    # 4. Download and extract labels
    label_tar = tmp / "PanTSMini_Label.tar.gz"
    download(PANTS_LABEL_URL, label_tar)
    label_extract_dir = tmp / "LabelAll"
    if not label_extract_dir.exists():
        label_extract_dir.mkdir(parents=True, exist_ok=True)
        print("  Extracting PanTSMini_Label.tar.gz ...")
        run(f'tar -xzf "{label_tar}" -C "{label_extract_dir}"')

    # 5. Stage cases
    gt_rows = []
    for idx, pants_id in enumerate(tumor_ids, start=41):
        cruz_id = f"Cruz_{idx:08d}"
        print(f"  [{idx-40}/20] {pants_id} -> {cruz_id}")

        # Find the CT image (PanTS stores as PanTS_XXXXXXXX/ct.nii.gz)
        img_src = img_extract_dir / pants_id / "ct.nii.gz"
        if not img_src.exists():
            # Try alternate naming
            for ext in ["CT.nii.gz", "ct.nii.gz", "image.nii.gz"]:
                alt = img_extract_dir / pants_id / ext
                if alt.exists():
                    img_src = alt
                    break

        if not img_src.exists():
            print(f"    WARNING: CT not found for {pants_id}, skipping")
            continue

        # Copy CT
        ct_dest = pub_dir / cruz_id / "ct.nii.gz"
        ct_dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(img_src), str(ct_dest))

        # Find label — PanTS labels may be in LabelAll/PanTS_XXXXXXXX/
        seg_src = label_extract_dir / pants_id / "label.nii.gz"
        if not seg_src.exists():
            for alt_name in ["seg.nii.gz", "segmentation.nii.gz", "label.nii.gz"]:
                alt = label_extract_dir / pants_id / alt_name
                if alt.exists():
                    seg_src = alt
                    break

        if seg_src.exists():
            # PanTS labels: 1=pancreas, 2=tumor (PDAC)
            has_lesion = extract_organ_lesion(
                seg_src, organ_label=1, tumor_label=2,
                out_dir=masks_dir / cruz_id
            )
        else:
            # If no label file, mark as tumor-positive from metadata
            print(f"    WARNING: Label not found for {pants_id}, marking from metadata")
            has_lesion = 1

        gt_rows.append(f"{cruz_id},pancreas,{has_lesion}")

    # Write ground truth CSV
    priv_dir.mkdir(parents=True, exist_ok=True)
    gt_path = priv_dir / "ground_truth.csv"
    with open(gt_path, "w") as f:
        f.write("patient_id,organ,lesion_present\n")
        f.write("\n".join(gt_rows) + "\n")
    print(f"\n  Wrote {gt_path}")


# ═══════════════════════════════════════════════════════════════════════════
# CLEANUP
# ═══════════════════════════════════════════════════════════════════════════
def cleanup():
    print("\n" + "=" * 60)
    print("CLEANING UP temp files")
    print("=" * 60)
    if TMP_DIR.exists():
        shutil.rmtree(str(TMP_DIR))
        print(f"  Removed {TMP_DIR}")


def print_summary():
    print("\n" + "=" * 60)
    print("DATA STAGING COMPLETE")
    print("=" * 60)
    for organ in ["Kidney", "Liver", "Pancreas"]:
        organ_dir = DATA_DIR / organ
        pub = organ_dir / "public"
        priv = organ_dir / "private"
        n_pub = len(list(pub.glob("Cruz_*/ct.nii.gz"))) if pub.exists() else 0
        gt = priv / "ground_truth.csv"
        print(f"\n  {organ}:")
        print(f"    Public CTs:    {n_pub}")
        print(f"    Ground truth:  {gt.exists()}")
        if gt.exists():
            with open(gt) as f:
                lines = f.readlines()[1:]  # skip header
                pos = sum(1 for l in lines if l.strip().endswith(",1"))
                print(f"    Tumor+:        {pos}/{len(lines)}")
        # List patients
        if pub.exists():
            patients = sorted([p.name for p in pub.iterdir() if p.is_dir()])
            print(f"    Patients:      {patients[0]} .. {patients[-1]}" if patients else "    Patients: none")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Stage benchmark data")
    parser.add_argument("--organ", choices=["kidney", "liver", "pancreas", "all"],
                        default="all", help="Which organ to stage")
    parser.add_argument("--no-cleanup", action="store_true",
                        help="Keep temp download files")
    args = parser.parse_args()

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    TMP_DIR.mkdir(parents=True, exist_ok=True)

    if args.organ in ("kidney", "all"):
        stage_kidney()
    if args.organ in ("liver", "all"):
        stage_liver()
    if args.organ in ("pancreas", "all"):
        stage_pancreas()

    if not args.no_cleanup:
        cleanup()

    print_summary()
